from flask import Blueprint, send_file, request, jsonify
from flask_login import login_required, current_user
from models import Project, Expense, Payment, ClientPayment
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import io

export_bp = Blueprint('export', __name__)

@export_bp.route('/daily-balance-pdf', methods=['POST'])
@login_required
def export_daily_balance_pdf():
    """Export daily balance sheet as PDF"""
    data = request.get_json()
    project_id = data.get('project_id')
    date_str = data.get('date')
    
    project = Project.query.filter_by(
        project_id=project_id,
        company_id=current_user.company_id
    ).first_or_404()
    
    target_date = datetime.fromisoformat(date_str).date()
    
    # Create PDF buffer
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e40af')
    )
    
    # Title
    elements.append(Paragraph(f"Daily Balance Sheet", title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Project details
    details = [
        ['Project:', project.name],
        ['Date:', target_date.strftime('%Y-%m-%d')],
        ['Company:', current_user.company.name]
    ]
    
    details_table = Table(details, colWidths=[2*inch, 4*inch])
    details_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.grey)
    ]))
    
    elements.append(details_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Calculate balances (simplified version)
    # In production, use the same logic as dashboard_api
    
    # Balance summary table
    balance_data = [
        ['Metric', 'Amount (₹)'],
        ['Opening Balance', '0.00'],
        ['Client Receipts (Today)', '0.00'],
        ['Payments Made (Today)', '0.00'],
        ['Expenses (Today)', '0.00'],
        ['Closing Balance', '0.00']
    ]
    
    balance_table = Table(balance_data, colWidths=[3*inch, 2*inch])
    balance_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(balance_table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'daily_balance_{project.name}_{date_str}.pdf',
        mimetype='application/pdf'
    )

@export_bp.route('/daily-balance-excel', methods=['POST'])
@login_required
def export_daily_balance_excel():
    """Export daily balance sheet as Excel"""
    data = request.get_json()
    project_id = data.get('project_id')
    date_str = data.get('date')
    
    project = Project.query.filter_by(
        project_id=project_id,
        company_id=current_user.company_id
    ).first_or_404()
    
    target_date = datetime.fromisoformat(date_str).date()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Balance"
    
    # Header styling
    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Title
    ws['A1'] = 'Daily Balance Sheet'
    ws['A1'].font = Font(bold=True, size=16, color="1e40af")
    
    # Project details
    ws['A3'] = 'Project:'
    ws['B3'] = project.name
    ws['A4'] = 'Date:'
    ws['B4'] = date_str
    ws['A5'] = 'Company:'
    ws['B5'] = current_user.company.name
    
    # Balance summary
    ws['A7'] = 'Metric'
    ws['B7'] = 'Amount (₹)'
    ws['A7'].fill = header_fill
    ws['B7'].fill = header_fill
    ws['A7'].font = header_font
    ws['B7'].font = header_font
    
    balance_rows = [
        ['Opening Balance', 0.00],
        ['Client Receipts (Today)', 0.00],
        ['Payments Made (Today)', 0.00],
        ['Expenses (Today)', 0.00],
        ['Closing Balance', 0.00]
    ]
    
    row = 8
    for metric, amount in balance_rows:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = amount
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'daily_balance_{project.name}_{date_str}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
