import openpyxl
import psycopg2
from datetime import datetime

# ── DB connection ──────────────────────────────────────────────
conn = psycopg2.connect(
    host="dpg-d7i7as5ckfvc73evmme0-a.oregon-postgres.render.com",
    dbname="cons_java",
    user="cons_java_user",
    password="LJWQRjMU2t42LPfJAyGqgT8Z97ukL4oq",
    port=5432
)
cur = conn.cursor()

# ── Constants ──────────────────────────────────────────────────
COMPANY_ID = 1

# ── Step 1: Add missing subcategory (Office Asset under TSK Expense id=11) ──
print("Step 1: Adding missing subcategory 'Office Asset'...")
cur.execute("""
    INSERT INTO sub_category (name, default_unit, parent_category_id)
    SELECT 'Office Asset', NULL, 11
    WHERE NOT EXISTS (
        SELECT 1 FROM sub_category WHERE name='Office Asset' AND parent_category_id=11
    );
""")
conn.commit()
print("  Done.")

# ── Step 2: Create Echanari project ───────────────────────────
print("\nStep 2: Creating 'Echanari' project...")
cur.execute("SELECT project_id FROM project WHERE name='Echanari' AND company_id=%s", (COMPANY_ID,))
row = cur.fetchone()
if row:
    project_id = row[0]
    print(f"  Project already exists with id={project_id}")
else:
    cur.execute("""
        INSERT INTO project (name, company_id, status, budget, created_at)
        VALUES (%s, %s, 'ACTIVE', 0, NOW())
        RETURNING project_id
    """, ('Echanari', COMPANY_ID))
    project_id = cur.fetchone()[0]
    conn.commit()
    print(f"  Created project id={project_id}")

# ── Step 3: Insert Client Payments ───────────────────────────
print("\nStep 3: Importing Payment Received sheet...")
wb = openpyxl.load_workbook(r"e:\Personal\Con_java\Echanari Expense.xlsx", data_only=True)
ws_pay = wb['Payment Received']

MODE_MAP = {
    'CASH': 'CASH',
    'UPI': 'UPI',
    'ACC': 'BANK',
    'BANK': 'BANK',
    'CHEQUE': 'BANK',
}

pay_count = 0
for i, row in enumerate(ws_pay.iter_rows(values_only=True)):
    # Skip header row (row 0 index 0)
    if i == 0:
        continue
    date_val, amount, paid_by, mode, description = row[0], row[1], row[2], row[3], row[4]
    if not date_val or not amount:
        continue
    if isinstance(date_val, datetime):
        pay_date = date_val.date()
    else:
        continue

    payment_mode = MODE_MAP.get(str(mode).strip().upper(), 'CASH') if mode else 'CASH'

    cur.execute("""
        INSERT INTO client_payment (company_id, project_id, amount, payment_date, payment_mode, remarks, created_at)
        VALUES (%s, %s, %s, %s, %s, %s, NOW())
    """, (COMPANY_ID, project_id, float(amount), pay_date, payment_mode, description))
    pay_count += 1
    print(f"  + {pay_date} | {payment_mode} | Rs.{amount:,.0f} | {description}")

conn.commit()
print(f"  Total client payments inserted: {pay_count}")

# ── Step 4: Insert Expenses ───────────────────────────────────
print("\nStep 4: Importing Expenses sheet...")
ws_exp = wb['Expenses']

# expense_type (col B) → (DB category name, DB subcategory name)
# DB schema: expense_type = subcategory, category = master category name
EXPENSE_MAP = {
    'MASON MATERIALS':   ('Material',    'Mason'),
    'CENTRING MATERIALS':('Material',    'Centring'),
    'MASON LABOUR':      ('Labour',      'Mason'),
    'CENTRING LABOUR':   ('Labour',      'Centring'),
    'CONCRETE GROUP':    ('Labour',      'Concrete Group'),
    'TEA':               ('Petty Cash',  'Tea'),
    'PETTY CASH':        ('Petty Cash',  'Others'),
    'DIESEL-MACHINE':    ('Petty Cash',  'Others'),
    'JCB':               ('JCB',         'JCB'),
    'CENTRING RENTALS':  ('Rental',      'Centring Rental'),
    'LABOUR TRANSPORT':  ('Transport',   'Labour'),
    'OTHER TRANSPORT':   ('Transport',   'Other'),
    'RENTAL TRANSPORT':  ('Transport',   'Other'),
    'SALARY':            ('TSK Expense', 'Salary'),
    'OTHERS':            ('OTHERS',      'Extra Work'),
    'OFFICE ASSET':      ('TSK Expense', 'Office Asset'),
    'OFFC ASSET':        ('TSK Expense', 'Office Asset'),
    # EXCLUDED:
    'BUILDING MATERIALS': None,
}

SKIP_TYPES = {'BUILDING MATERIALS'}

exp_count = 0
skip_count = 0
unknown_types = set()

for i, row in enumerate(ws_exp.iter_rows(values_only=True)):
    if i == 0:
        continue  # skip header
    date_val, exp_type, amount, description = row[0], row[1], row[2], row[3]

    if not date_val or not amount or not exp_type:
        continue
    if not isinstance(date_val, datetime):
        continue

    exp_type_clean = str(exp_type).strip().upper()

    # Skip excluded types
    if exp_type_clean in SKIP_TYPES:
        skip_count += 1
        continue

    mapping = EXPENSE_MAP.get(exp_type_clean)
    if mapping is None and exp_type_clean not in EXPENSE_MAP:
        unknown_types.add(exp_type_clean)
        continue

    category_name, subcategory_name = mapping
    exp_date = date_val.date()

    # expense_type = subcategory name, category = master category name
    cur.execute("""
        INSERT INTO expense (company_id, project_id, expense_type, category, amount, payment_mode, expense_date, description, created_at)
        VALUES (%s, %s, %s, %s, %s, 'CASH', %s, %s, NOW())
    """, (COMPANY_ID, project_id, subcategory_name, category_name, float(amount), exp_date, description))
    exp_count += 1

conn.commit()

print(f"  Total expenses inserted: {exp_count}")
print(f"  Skipped (BUILDING MATERIALS): {skip_count}")
if unknown_types:
    print(f"  ⚠ Unknown types NOT imported: {unknown_types}")
else:
    print("  All types mapped successfully.")

cur.close()
conn.close()
print("\n✅ Import complete!")
