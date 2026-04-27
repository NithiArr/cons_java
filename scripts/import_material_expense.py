import openpyxl
import psycopg2
from datetime import datetime

# ── DB connection ──────────────────────────────────────────────
conn = psycopg2.connect(
    host="dpg-d7i7as5ckfvc73evmme0-a.oregon-postgres.render.com",
    dbname="cons_java",
    user="cons_java_user",
    password="LJWQRjMU2t42LPfJAyGqgT8Z97ukL4oq",
    port=5432
)
cur = conn.cursor()

COMPANY_ID = 1
PROJECT_ID = 1  # We know Echanari project has ID 1

wb = openpyxl.load_workbook(r"e:\Personal\Con_java\Echanari Material Expense.xlsx", data_only=True)

# ── Helper mapping for Categories ─────────────────────────────
CATEGORY_MAP = {
    'STEEL': ('Steel', 'Steel'),
    'M-SAND': ('Building Materials', 'M-Sand'),
    '40 MM METAL': ('Building Materials', '40 MM'),
    'CEMENT': ('Building Materials', 'Cement'),
    '20 MM JALLY': ('Building Materials', '20 MM'),
    'BRICK-FLYASH': ('Brick', 'Fly Ash'),
    'BRICK': ('Brick', 'Brick'),
    'P-SAND': ('Building Materials', 'P-Sand')
}

# ── 1. Vendor Sync ─────────────────────────────────────────────
print("Step 1: Syncing Vendors...")
ws_vendors = wb['Vendor Balance Sheet']

# Load existing vendors
cur.execute("SELECT vendor_id, name FROM vendor WHERE company_id = %s", (COMPANY_ID,))
existing_vendors = {row[1].strip().upper(): row[0] for row in cur.fetchall()}

vendor_count = 0
for i, row in enumerate(ws_vendors.iter_rows(values_only=True)):
    if i == 0: continue
    v_name = row[0]
    if not v_name or v_name == 0: continue
    
    v_name_clean = str(v_name).strip()
    v_name_upper = v_name_clean.upper()
    
    if v_name_upper not in existing_vendors:
        cur.execute("""
            INSERT INTO vendor (name, company_id, created_at)
            VALUES (%s, %s, NOW()) RETURNING vendor_id
        """, (v_name_clean, COMPANY_ID))
        new_v_id = cur.fetchone()[0]
        existing_vendors[v_name_upper] = new_v_id
        vendor_count += 1
        print(f"  + Added new vendor: {v_name_clean}")

conn.commit()
print(f"  Total new vendors added: {vendor_count}")


# ── 2. Material Details ────────────────────────────────────────
print("\nStep 2: Importing Material Purchases...")
ws_materials = wb['Material Details']

mat_count = 0
for i, row in enumerate(ws_materials.iter_rows(values_only=True)):
    if i == 0: continue
    date_val, cat, company_brand, dealer, qty, unit, rate, total_bill = row[:8]
    
    if not date_val or not total_bill: continue
    if not isinstance(date_val, datetime): continue
        
    exp_date = date_val.date()
    cat_clean = str(cat).strip().upper() if cat else "UNKNOWN"
    brand = str(company_brand).strip() if company_brand else ""
    dealer_clean = str(dealer).strip().upper() if dealer else ""
    vendor_id = existing_vendors.get(dealer_clean)
    
    if vendor_id is None:
        print(f"  ⚠ Warning: Vendor '{dealer}' not found for row {i+1}. Skipping.")
        continue
        
    # Get mapping
    master_cat, item_name = CATEGORY_MAP.get(cat_clean, ('Building Materials', cat_clean))
    
    # Defaults for quantities
    q = float(qty) if qty else 1.0
    u = str(unit).strip() if unit else "unit"
    r = float(rate) if rate else float(total_bill) / q
    
    # Insert Expense
    cur.execute("""
        INSERT INTO expense (company_id, project_id, vendor_id, expense_type, category, amount, payment_mode, expense_date, created_at)
        VALUES (%s, %s, %s, 'Material Purchase', %s, %s, 'CREDIT', %s, NOW()) RETURNING expense_id
    """, (COMPANY_ID, PROJECT_ID, vendor_id, master_cat, float(total_bill), exp_date))
    expense_id = cur.fetchone()[0]
    
    # Insert Expense Item
    cur.execute("""
        INSERT INTO expense_item (expense_id, item_name, brand, quantity, measuring_unit, unit_price, total_price)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (expense_id, item_name, brand, q, u, r, float(total_bill)))
    
    mat_count += 1

conn.commit()
print(f"  Total material purchases inserted: {mat_count}")


# ── 3. Payment Details ─────────────────────────────────────────
print("\nStep 3: Importing Vendor Payments...")
ws_payments = wb['Payment Details']

MODE_MAP = {
    'CASH': 'CASH',
    'ACC': 'BANK',
    'UPI': 'UPI',
    'BANK': 'BANK'
}

pay_count = 0
for i, row in enumerate(ws_payments.iter_rows(values_only=True)):
    if i == 0: continue
    date_val, vendor_name, amount_paid, mode = row[:4]
    
    if not date_val or not amount_paid: continue
    if not isinstance(date_val, datetime): continue
        
    pay_date = date_val.date()
    v_name_clean = str(vendor_name).strip().upper() if vendor_name else ""
    vendor_id = existing_vendors.get(v_name_clean)
    
    if vendor_id is None:
        print(f"  ⚠ Warning: Vendor '{vendor_name}' not found for payment. Skipping.")
        continue
        
    p_mode = MODE_MAP.get(str(mode).strip().upper(), 'CASH') if mode else 'CASH'
    
    # Insert Payment
    cur.execute("""
        INSERT INTO payment (company_id, project_id, vendor_id, amount, payment_date, payment_mode, created_at)
        VALUES (%s, %s, %s, %s, %s, %s, NOW())
    """, (COMPANY_ID, PROJECT_ID, vendor_id, float(amount_paid), pay_date, p_mode))
    
    pay_count += 1

conn.commit()
print(f"  Total vendor payments inserted: {pay_count}")

cur.close()
conn.close()
print("\nImport complete!")
