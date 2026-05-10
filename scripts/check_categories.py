import openpyxl
import psycopg2

# ── 1. Pull all unique expense types from all 5 Expense sheets ──────────
files = {
    'DSP':        'Data/DSP Site Expense.xlsx',
    'Ganapathy':  'Data/Ganapathy Residence Expense.xlsx',
    'HappyTown':  'Data/Happy Town Expense.xlsx',
    'SK':         'Data/SK Expense Sheet.xlsx',
    'Velavan':    'Data/Velavan Clinic Expense.xlsx',
}

all_exp_types = set()
for proj, path in files.items():
    wb = openpyxl.load_workbook(path, data_only=True)
    for sh in ['Expenses', 'Expense']:
        if sh in wb.sheetnames:
            ws = wb[sh]
            for row in ws.iter_rows(values_only=True):
                if row[0] and hasattr(row[0], 'year') and row[1]:
                    all_exp_types.add(str(row[1]).strip().upper())

# Remove BUILDING MATERIALS (will be skipped — double entry from material sheet)
all_exp_types.discard('BUILDING MATERIALS')

# ── 2. Pull all unique material categories from Material Details sheets ──
mat_files = {
    'DSP':       'Data/DSP Material Expense.xlsx',
    'Ganapathy': 'Data/Ganapathy Residence Material Expense.xlsx',
    'HappyTown': 'Data/Happy Town Material Expense.xlsx',
    'SK':        'Data/SK Material Expense.xlsx',
    'Velavan':   'Data/Velavan Material Expense.xlsx',
}

all_mat_types = set()
for proj, path in mat_files.items():
    wb = openpyxl.load_workbook(path, data_only=True)
    if 'Material Details' in wb.sheetnames:
        ws = wb['Material Details']
        for row in ws.iter_rows(values_only=True):
            if row[0] and hasattr(row[0], 'year') and row[1]:
                all_mat_types.add(str(row[1]).strip().upper())

# ── 3. Connect to Render DB and get existing categories + subcategories ──
conn = psycopg2.connect(
    host='dpg-d7i7as5ckfvc73evmme0-a.oregon-postgres.render.com',
    port=5432, dbname='cons_java',
    user='cons_java_user',
    password='LJWQRjMU2t42LPfJAyGqgT8Z97ukL4oq',
    sslmode='require'
)
cur = conn.cursor()

cur.execute("SELECT category_id, name, type FROM master_category")
db_cats = {r[1].upper(): (r[0], r[2]) for r in cur.fetchall()}

cur.execute("""SELECT sc.name, mc.name, mc.type
               FROM sub_category sc JOIN master_category mc ON mc.category_id = sc.parent_category_id""")
db_subcats = {(r[1].upper(), r[0].upper()): r[2] for r in cur.fetchall()}  # key=(cat_name, subcat_name)

cur.close(); conn.close()

# ── 4. Print results ─────────────────────────────────────────────────────
print("=" * 60)
print("EXPENSE TYPES NOT MATCHING DB (needs new cat/subcat)")
print("=" * 60)
print("(BUILDING MATERIALS omitted — will be skipped in import)\n")

# Known mappings (excel type → db category)
known_expense_map = {
    'PETTY CASH': 'PETTY CASH',
    'MASON LABOUR': 'LABOUR',
    'CENTRING LABOUR': 'LABOUR',
    'MASON MATERIALS': 'MATERIAL',
    'CENTRING MATERIALS': 'MATERIAL',
    'CONCRETE GROUP': 'LABOUR',
    'CENTRING RENTALS': 'RENTAL',
    'JCB': 'JCB',
    'OTHER TRANSPORT': 'TRANSPORT',
    'LABOUR TRANSPORT': 'TRANSPORT',
    'TEA': 'PETTY CASH',
    'SALARY': 'TSK EXPENSE',
    'ELECTRICAL LABOUR': 'LABOUR',
    'ELECTRICIAL LABOUR': 'LABOUR',
    'ELECTRICIAL': 'ELECTRICAL',
    'ELECTRICAL MATERIALS': 'ELECTRICAL',
    'ELECTRICAL LABOUR': 'LABOUR',
    'PAINTING LABOUR': 'LABOUR',
    'PAINTING MATERIALS': 'PAINTING',
    'PLUMBING MATERIALS': 'PLUMBING',
    'OTHERS': 'OTHERS',
    'OFFC ASSET': 'TSK EXPENSE',
    'OFFICE ASSET': 'TSK EXPENSE',
    'RO WATER': 'PETTY CASH',
    'SALT WATER': 'PETTY CASH',
    'DIESEL': 'PETTY CASH',
    'DIESEL-MACHINE': 'PETTY CASH',
}

missing = []
for etype in sorted(all_exp_types):
    db_cat = known_expense_map.get(etype)
    if db_cat is None:
        missing.append((etype, 'NO MAPPING DEFINED'))
    elif db_cat not in db_cats:
        missing.append((etype, f'Category "{db_cat}" NOT IN DB'))

print("--- Expense types with NO DB category mapping at all ---")
for item in missing:
    print(f"  ❌  {item[0]}")

# Now check subcategories that need to be added
print()
print("--- Expense types that need new SUBCATEGORIES in DB ---")
needs_subcat = {
    'CARPENTER LABOUR':    ('LABOUR', 'Carpenter'),
    'TILES LABOUR':        ('LABOUR', 'Tiler'),
    'DIESEL':              ('PETTY CASH', 'Diesel'),
    'DIESEL-MACHINE':      ('PETTY CASH', 'Diesel Machine'),
    'RO WATER':            ('PETTY CASH', 'RO Water'),
    'SALT WATER':          ('PETTY CASH', 'Salt Water'),
    'DESIGN':              ('OTHERS', 'Design'),
    'DESIGN CHARGES':      ('OTHERS', 'Design'),
    'FABRICATON WORK':     ('OTHERS', 'Fabrication Work'),
    'FABRICATION WORK':    ('OTHERS', 'Fabrication Work'),
    'JOINERIES':           ('OTHERS', 'Joineries'),
    'TILES PURCHASE':      ('OTHERS', 'Tiles Purchase'),
    'PILE - COMPOUND':     ('OTHERS', 'Pile / Compound'),
    'PILE-COMPOUND':       ('OTHERS', 'Pile / Compound'),
    'DIGITAL SURVEY':      ('OTHERS', 'Digital Survey'),
    'GSQUARE SITE':        ('OTHERS', 'Inter-Site Transfer'),
    'SK SITE':             ('OTHERS', 'Inter-Site Transfer'),
}

printed = set()
for etype in sorted(all_exp_types):
    if etype in needs_subcat:
        parent_cat, subcat = needs_subcat[etype]
        key = (parent_cat, subcat)
        if key not in printed:
            in_db = (parent_cat, subcat.upper()) in db_subcats
            status = '✅ already in DB' if in_db else '❌ MISSING'
            print(f"  {status}  Subcat: \"{subcat}\" under Category: \"{parent_cat}\"")
            printed.add(key)

print()
print("=" * 60)
print("MATERIAL CATEGORIES NOT MATCHING DB")
print("=" * 60)

known_mat_map = {
    'STEEL':        'STEEL',
    'CEMENT':       'BUILDING MATERIALS',
    'M-SAND':       'BUILDING MATERIALS',
    'P-SAND':       'BUILDING MATERIALS',
    '20 MM JALLY':  'BUILDING MATERIALS',
    '40 MM METAL':  'BUILDING MATERIALS',
    'GRAVEL':       'BUILDING MATERIALS',
    'BABY CHIPS':   'BUILDING MATERIALS',
    'BRICK':        'BRICK',
    'BRICK-RED':    'BRICK',
    'BRICK-FLYASH': 'BRICK',
    'ELECTRICAL':   'ELECTRICAL',
    'PLUMBING':     'PLUMBING',
    'PAINTING':     'PAINTING',
    'SOLID BLOCKS': 'BRICK',
    'OTHERS':       'OTHERS',
    'TILES':        None,    # no category
}

for mtype in sorted(all_mat_types):
    db_cat = known_mat_map.get(mtype)
    if db_cat is None:
        print(f"  ❌  \"{mtype}\" — NO DB category (needs new category or mapping)")
    elif db_cat not in db_cats:
        print(f"  ❌  \"{mtype}\" → wants cat \"{db_cat}\" — NOT IN DB")
    else:
        pass  # matched

print()
print("All expense types found (after removing BUILDING MATERIALS):")
for t in sorted(all_exp_types):
    print(f"  {t}")
print()
print("All material categories found:")
for t in sorted(all_mat_types):
    print(f"  {t}")
