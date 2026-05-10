import openpyxl
import psycopg2
from datetime import datetime

print("Connecting to database...")
conn = psycopg2.connect(
    host='dpg-d7i7as5ckfvc73evmme0-a.oregon-postgres.render.com',
    port=5432, dbname='cons_java',
    user='cons_java_user',
    password='LJWQRjMU2t42LPfJAyGqgT8Z97ukL4oq',
    sslmode='require'
)
cur = conn.cursor()

COMPANY_ID = 1

def execute(q, vars=None):
    cur.execute(q, vars)

# ── 1. ADD MASTER CATEGORIES ──────────────────────────────────────────────────
print("Setting up Master Categories...")
new_masters = [
    ('Fuel', 'EXPENSE'),
    ('Design', 'EXPENSE'),
    ('Digital Survey', 'EXPENSE'),
    ('Fabrication', 'EXPENSE'),
    ('Joineries', 'EXPENSE'),
    ('Pile / Compound', 'EXPENSE'),
    ('Site Transfer', 'EXPENSE'),
    ('Tiles', 'MATERIAL')
]

for name, mtype in new_masters:
    execute("SELECT category_id FROM master_category WHERE name=%s AND type=%s", (name, mtype))
    if not cur.fetchone():
        execute("INSERT INTO master_category (name, type, is_active) VALUES (%s, %s, true)", (name, mtype))

execute("SELECT UPPER(name), category_id FROM master_category")
mc_map = {r[0]: r[1] for r in cur.fetchall()}

# ── 2. ADD SUBCATEGORIES ──────────────────────────────────────────────────────
print("Setting up Sub Categories...")
new_subcats = [
    ('Labour', 'Carpenter', ''),
    ('Labour', 'Tiles', ''),
    ('Petty Cash', 'RO Water', ''),
    ('Petty Cash', 'Salt Water', ''),
    ('Petty Cash', 'Diesel', ''),
    ('Petty Cash', 'Diesel Machine', ''),
    ('Fuel', 'Diesel', ''),
    ('Fuel', 'Petrol', ''),
    ('Design', 'Charges', ''),
    ('Digital Survey', 'Charges', ''),
    ('Fabrication', 'Charges', ''),
    ('Joineries', 'Charges', ''),
    ('Pile / Compound', 'Pile / Compound', ''),
    ('Site Transfer', 'Internal', ''),
    ('Tiles', 'Tiles', ''),
    ('Brick', 'Solid Block', 'nos')
]

for p_name, s_name, unit in new_subcats:
    p_id = mc_map.get(p_name.upper())
    if p_id:
        execute("SELECT subcategory_id FROM sub_category WHERE name=%s AND parent_category_id=%s", (s_name, p_id))
        if not cur.fetchone():
            execute("INSERT INTO sub_category (name, default_unit, parent_category_id) VALUES (%s, %s, %s)", (s_name, unit, p_id))

# ── 3. ADD VENDORS ────────────────────────────────────────────────────────────
print("Setting up Vendors...")
new_vendors = [
    'ANJENEYA', 'NV ASSOCIATES', 'SKANDA STEELS', 'RAMDEV HARDWARES',
    'RAMDEV ELECTRICALS', 'SALEM FLYASH', 'SUSSHANT', 'VASANTHAM AGENCIES',
    'LAKSHMI CERAMICS', 'SS TRADERS', 'LOCAL'
]

for v in new_vendors:
    execute("SELECT vendor_id FROM vendor WHERE name=%s AND company_id=%s", (v, COMPANY_ID))
    if not cur.fetchone():
        execute("INSERT INTO vendor (name, company_id, created_at) VALUES (%s, %s, NOW())", (v, COMPANY_ID))

execute("SELECT UPPER(name), vendor_id FROM vendor WHERE company_id=%s", (COMPANY_ID,))
db_vendors = {r[0]: r[1] for r in cur.fetchall()}

# Deduplication mappings
if 'VEL BRICKS' in db_vendors:
    db_vendors['VEL BRICK'] = db_vendors['VEL BRICKS']
if 'AMARUL FLYASH' in db_vendors:
    db_vendors['AMARUN FLYASH'] = db_vendors['AMARUL FLYASH']

# ── 4. ADD PROJECTS ───────────────────────────────────────────────────────────
print("Setting up Projects...")
projects_to_add = ['DSP', 'Ganapathy Residence', 'Happy Town', 'SK', 'Velavan Clinic']
for p in projects_to_add:
    execute("SELECT project_id FROM project WHERE name=%s AND company_id=%s", (p, COMPANY_ID))
    if not cur.fetchone():
        execute("INSERT INTO project (name, status, company_id, budget, created_at) VALUES (%s, 'ACTIVE', %s, 0, NOW())", (p, COMPANY_ID))

execute("SELECT name, project_id FROM project WHERE company_id=%s", (COMPANY_ID,))
db_projects = {r[0]: r[1] for r in cur.fetchall()}

# ── 5. CLEANUP EXISTING DATA FOR THESE PROJECTS TO BE IDEMPOTENT ──────────────
p_ids = tuple(db_projects[p] for p in projects_to_add)
if p_ids:
    print("Clearing old data for these 5 projects (if any) to prevent duplication...")
    execute("DELETE FROM client_payment WHERE project_id IN %s", (p_ids,))
    execute("DELETE FROM payment WHERE project_id IN %s", (p_ids,))
    execute("DELETE FROM expense_item WHERE expense_id IN (SELECT expense_id FROM expense WHERE project_id IN %s)", (p_ids,))
    execute("DELETE FROM expense WHERE project_id IN %s", (p_ids,))

# ── MAPPING FUNCTIONS ─────────────────────────────────────────────────────────

def map_expense_category(excel_type):
    mapping = {
        'PETTY CASH': ('Petty Cash', 'Others'),
        'TEA': ('Petty Cash', 'Tea'),
        'MASON LABOUR': ('Labour', 'Mason'),
        'CENTRING LABOUR': ('Labour', 'Centring'),
        'ELECTRICAL LABOUR': ('Labour', 'Electrician'),
        'ELECTRICIAL LABOUR': ('Labour', 'Electrician'),
        'PAINTING LABOUR': ('Labour', 'Painter'),
        'CONCRETE GROUP': ('Labour', 'Concrete Group'),
        'MASON MATERIALS': ('Material', 'Mason'),
        'CENTRING MATERIALS': ('Material', 'Centring'),
        'CENTRING RENTALS': ('Rental', 'Centring Rental'),
        'JCB': ('JCB', 'JCB'),
        'LABOUR TRANSPORT': ('Transport', 'Labour'),
        'OTHER TRANSPORT': ('Transport', 'Other'),
        'SALARY': ('TSK Expense', 'Salary'),
        'OFFC ASSET': ('TSK Expense', 'Office Asset'),
        'OFFICE ASSET': ('TSK Expense', 'Office Asset'),
        'ELECTRICAL MATERIALS': ('Electrical', 'Wires'),
        'ELECTRICIAL': ('Electrical', 'Wires'),
        'PAINTING MATERIALS': ('Painting', 'Other'),
        'PLUMBING MATERIALS': ('Plumbing', 'Other'),
        'OTHERS': ('OTHERS', 'Extra Work'),
        'DIESEL': ('Fuel', 'Diesel'),
        'DIESEL-MACHINE': ('Fuel', 'Diesel'),
        'PETROL-GENSET': ('Fuel', 'Petrol'),
        'CARPENTER LABOUR': ('Labour', 'Carpenter'),
        'TILES LABOUR': ('Labour', 'Tiles'),
        'RO WATER': ('Petty Cash', 'RO Water'),
        'SALT WATER': ('Petty Cash', 'Salt Water'),
        'DESIGN': ('Design', 'Charges'),
        'DESIGN CHARGES': ('Design', 'Charges'),
        'DIGITAL SURVEY': ('Digital Survey', 'Charges'),
        'FABRICATON WORK': ('Fabrication', 'Charges'),
        'FABRICATION WORK': ('Fabrication', 'Charges'),
        'JOINERIES': ('Joineries', 'Charges'),
        'PILE - COMPOUND': ('Pile / Compound', 'Pile / Compound'),
        'PILE-COMPOUND': ('Pile / Compound', 'Pile / Compound'),
        'GSQUARE SITE': ('Site Transfer', 'Internal'),
        'SK SITE': ('Site Transfer', 'Internal'),
    }
    return mapping.get(excel_type)

def map_material_category(excel_type):
    mapping = {
        'STEEL': ('Steel', None),
        'CEMENT': ('Building Materials', 'Cement'),
        'M-SAND': ('Building Materials', 'M-Sand'),
        'P-SAND': ('Building Materials', 'P-Sand'),
        '20 MM JALLY': ('Building Materials', '20 MM'),
        '40 MM METAL': ('Building Materials', '40 MM'),
        'GRAVEL': ('Building Materials', 'Gravel'),
        'BABY CHIPS': ('Building Materials', 'Baby Chips'),
        'BRICK': ('Brick', 'Red'),
        'BRICK-RED': ('Brick', 'Red'),
        'BRICK-FLYASH': ('Brick', 'Fly Ash'),
        'SOLID BLOCKS': ('Brick', 'Solid Block'),
        'ELECTRICAL': ('Electrical', 'Wires'),
        'PLUMBING': ('Plumbing', 'Fitting'),
        'PAINTING': ('Painting', 'Other'),
        'TILES': ('Tiles', 'Tiles'),
        'OTHERS': ('OTHERS', 'Extra Work'),
    }
    return mapping.get(excel_type)

def map_payment_mode(mode):
    if not mode: return 'CASH'
    m = str(mode).strip().upper()
    if m == 'ACC': return 'BANK'
    if m in ['CASH', 'BANK', 'UPI', 'CREDIT']: return m
    return 'CASH'

# ── 6. LOAD DATA ──────────────────────────────────────────────────────────────

files = {
    'DSP':        ('Data/DSP Site Expense.xlsx', 'Data/DSP Material Expense.xlsx'),
    'Ganapathy Residence':  ('Data/Ganapathy Residence Expense.xlsx', 'Data/Ganapathy Residence Material Expense.xlsx'),
    'Happy Town':  ('Data/Happy Town Expense.xlsx', 'Data/Happy Town Material Expense.xlsx'),
    'SK':         ('Data/SK Expense Sheet.xlsx', 'Data/SK Material Expense.xlsx'),
    'Velavan Clinic':    ('Data/Velavan Clinic Expense.xlsx', 'Data/Velavan Material Expense.xlsx'),
}

for proj_name, (exp_file, mat_file) in files.items():
    project_id = db_projects[proj_name]
    print(f"\n---> Loading Project: {proj_name}")

    # --- EXPENSE FILE ---
    print("     Reading Expenses sheet...")
    wb_exp = openpyxl.load_workbook(exp_file, data_only=True)

    # Client Payments
    if 'Payment Received' in wb_exp.sheetnames:
        ws = wb_exp['Payment Received']
        cp_count = 0
        for row in ws.iter_rows(values_only=True):
            if row[0] and hasattr(row[0], 'year') and isinstance(row[1], (int,float)):
                dt = row[0].strftime('%Y-%m-%d')
                amt = row[1]
                mode = map_payment_mode(row[3])
                desc = str(row[5]) if len(row) > 5 and row[5] else None
                execute("INSERT INTO client_payment (company_id, project_id, amount, payment_date, payment_mode, remarks, created_at) VALUES (%s, %s, %s, %s, %s, %s, NOW())",
                        (COMPANY_ID, project_id, amt, dt, mode, desc))
                cp_count += 1
        print(f"       + {cp_count} Client Payments")

    # Regular Expenses & "TILES PURCHASE"
    tiles_purchases = []
    exp_count = 0
    for sh in ['Expenses', 'Expense']:
        if sh in wb_exp.sheetnames:
            ws = wb_exp[sh]
            for row in ws.iter_rows(values_only=True):
                if row[0] and hasattr(row[0], 'year') and row[1] and isinstance(row[2], (int,float)):
                    dt = row[0].strftime('%Y-%m-%d')
                    etype_raw = str(row[1]).strip().upper()
                    amt = row[2]
                    desc = str(row[3]) if len(row) > 3 and row[3] else None

                    if etype_raw == 'BUILDING MATERIALS':
                        continue # skip
                    if etype_raw == 'TILES PURCHASE':
                        tiles_purchases.append({
                            'date': dt, 'amount': amt, 'desc': desc
                        })
                        continue

                    cat_map = map_expense_category(etype_raw)
                    cat_str = cat_map[0] if cat_map else etype_raw

                    execute("INSERT INTO expense (company_id, project_id, expense_type, expense_date, category, amount, payment_mode, description, created_at) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW())",
                            (COMPANY_ID, project_id, 'Regular Expense', dt, cat_str, amt, 'CASH', desc))
                    exp_count += 1
            print(f"       + {exp_count} Regular Expenses")

    # --- MATERIAL FILE ---
    print("     Reading Material sheet...")
    wb_mat = openpyxl.load_workbook(mat_file, data_only=True)

    # Vendor Payments
    vp_count = 0
    if 'Payment Details' in wb_mat.sheetnames:
        ws = wb_mat['Payment Details']
        for row in ws.iter_rows(values_only=True):
            if row[0] and hasattr(row[0], 'year') and row[1] and isinstance(row[2], (int,float)):
                dt = row[0].strftime('%Y-%m-%d')
                v_name = str(row[1]).strip().upper()
                vendor_id = db_vendors.get(v_name)
                amt = row[2]
                mode = map_payment_mode(row[3])
                if vendor_id:
                    execute("INSERT INTO payment (company_id, project_id, vendor_id, amount, payment_date, payment_mode, created_at) VALUES (%s, %s, %s, %s, %s, %s, NOW())",
                            (COMPANY_ID, project_id, vendor_id, amt, dt, mode))
                    vp_count += 1
        print(f"       + {vp_count} Vendor Payments")

    # Material Purchases
    mat_count = 0
    if 'Material Details' in wb_mat.sheetnames:
        ws = wb_mat['Material Details']
        for row in ws.iter_rows(values_only=True):
            if row[0] and hasattr(row[0], 'year') and row[1] and isinstance(row[7], (int,float)):
                dt = row[0].strftime('%Y-%m-%d')
                cat_raw = str(row[1]).strip().upper()
                brand = str(row[2]) if row[2] else None
                v_name = str(row[3]).strip().upper() if row[3] else None
                vendor_id = db_vendors.get(v_name) if v_name else None
                qty = row[4] if isinstance(row[4], (int,float)) else 1.0
                unit = str(row[5]) if row[5] else 'Unit'
                rate = row[6]
                total = row[7]

                if not isinstance(rate, (int,float)) or rate == 0:
                    rate = total / qty if qty else 0

                cat_map = map_material_category(cat_raw)
                cat_str = cat_map[0] if cat_map else cat_raw
                item_name = cat_map[1] if cat_map and cat_map[1] else cat_raw

                # Insert expense
                execute("INSERT INTO expense (company_id, project_id, vendor_id, expense_type, expense_date, category, amount, payment_mode, created_at) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW()) RETURNING expense_id",
                        (COMPANY_ID, project_id, vendor_id, 'Material Purchase', dt, cat_str, total, 'CASH'))
                expense_id = cur.fetchone()[0]

                # Insert expense item
                execute("INSERT INTO expense_item (expense_id, item_name, quantity, measuring_unit, unit_price, total_price, brand) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                        (expense_id, item_name, qty, unit, rate, total, brand))
                mat_count += 1

    # Handle TILES PURCHASE from expense sheet as Material Purchase
    for tp in tiles_purchases:
        execute("INSERT INTO expense (company_id, project_id, expense_type, expense_date, category, amount, payment_mode, description, created_at) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW()) RETURNING expense_id",
                (COMPANY_ID, project_id, 'Material Purchase', tp['date'], 'Tiles', tp['amount'], 'CASH', tp['desc']))
        expense_id = cur.fetchone()[0]
        execute("INSERT INTO expense_item (expense_id, item_name, quantity, measuring_unit, unit_price, total_price) VALUES (%s, %s, %s, %s, %s, %s)",
                (expense_id, 'Tiles', 1.0, 'Unit', tp['amount'], tp['amount']))
        mat_count += 1
        
    print(f"       + {mat_count} Material Purchases")

conn.commit()
cur.close()
conn.close()
print("\n🎉 ALL DATA PUSHED SUCCESSFULLY to Render DB!")
