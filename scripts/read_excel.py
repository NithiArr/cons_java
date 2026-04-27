import openpyxl

wb = openpyxl.load_workbook(r"e:\Personal\Con_java\Echanari Expense.xlsx", data_only=True)
print(f"Sheets: {wb.sheetnames}")

# Print ALL rows from the main expense sheet and payment received sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    print(f"\n=== Sheet: '{sheet_name}' ===")
    for i, row in enumerate(rows[:5]):
        print(f"  Header Row {i+1}: {list(row)}")
    break  # Just show first sheet headers for now

# Focus on payment received sheet
for sn in wb.sheetnames:
    if 'payment' in sn.lower() or 'received' in sn.lower() or 'client' in sn.lower():
        ws = wb[sn]
        print(f"\n=== PAYMENT SHEET: '{sn}' ===")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if any(v is not None for v in row):
                print(f"  Row {i+1}: {list(row)}")
